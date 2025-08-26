import openpyxl
from pathlib import Path
from robot.libraries.BuiltIn import BuiltIn
import os

class excel_library:

    #def __init__(self):
        #self.workbook = None

    def read_cell(self, sheet_name, col_name):
        current_directory = os.getcwd()
        project_dir = Path(__file__).resolve().parent.parent
        print("Project Directory:", project_dir)  # Debugging print
        #scenarios_path = project_dir / 'dataFiles' / 'scenarios.xlsx'
        #scenarios_path = "./Template/datafiles/scenarios.xlsx"

        if os.name == 'nt':  # Windows
            scenarios_path = project_dir / 'datafiles' / 'scenarios.xlsx'
        else:  # Linux/Mac
            scenarios_path = "./Template/datafiles/scenarios.xlsx"
        print("Scenarios File Path:", scenarios_path)  # Debugging print
        self.workbook = openpyxl.load_workbook(scenarios_path)
        sheet = self.workbook[sheet_name]
        row = self.get_row_num(sheet_name)
        col = self.get_col_num(sheet_name, col_name)
        return sheet.cell(int(row), int(col)).value

    def write_cell(self, sheet_name, col_name, value):
        current_directory = os.getcwd()
        project_dir = os.path.dirname(os.path.dirname(__file__))
        #scenarios_path = os.path.join(project_dir, 'datafiles', 'scenarios.xlsx')
        #scenarios_path = "./Template/datafiles/scenarios.xlsx"
        if os.name == 'nt':  # Windows
            scenarios_path = os.path.join(project_dir, 'datafiles', 'scenarios.xlsx')
        else:  # Linux/Mac
            scenarios_path = "./Template/datafiles/scenarios.xlsx"
        self.workbook = openpyxl.load_workbook(scenarios_path)
        sheet = self.workbook[sheet_name]
        row = self.get_row_num(sheet_name)
        col = self.get_col_num(sheet_name, col_name)
        sheet.cell(int(row), int(col), value)
        self.workbook.save(scenarios_path)

    def get_row_num(self, sheet_name):

        built_in = BuiltIn()
        test_name = built_in.get_variable_value('${TEST_NAME}')
        sheet = self.workbook[sheet_name]
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            for cell in row:
                if cell.value == str(test_name):
                    Row = int(cell.row)
                    break
        return Row

    def get_col_num(self, sheet_name, col_name):

        sheet = self.workbook[sheet_name]
        for col in sheet.iter_cols(min_col=None, max_col=sheet.max_column):
            for cell in col:
                if cell.value == str(col_name):
                    column = int (cell.column)
                    break
        return column


