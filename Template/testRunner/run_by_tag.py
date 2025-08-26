import os
import subprocess
import time
from pathlib import Path
from openpyxl import load_workbook
import sys

os_name = os.name

def run_SelectedTest(tag):      
    # Dynamically determine the project root directory
    project_root = Path(__file__).resolve().parent.parent

    # Construct the file path for scenarios.xlsx
    filepath = project_root / 'datafiles' / 'scenarios.xlsx'

    workbook = load_workbook(filepath)
    sheet = workbook['Scenario']

    scenario_list = []
    modified_list = []

    for row in sheet.iter_rows(min_row=None, values_only=True):  # Skipping the header row
        Scenerio, Run_Status = row  # Unpack the row values
        if Run_Status and str(Run_Status).lower() == 'yes':  # Check if status is 'yes'
            scenario_list.append(Scenerio)  # Add to the list

    if not scenario_list: #check my list has any value
        print("No Scenerio marked to run") 

    for file in scenario_list:
            modified_list.append(file + '.robot')
           
    for file in modified_list:
            if file.endswith('.robot'):
                # Construct the test file path
                test_file = project_root / 'tests' / file
                print(f'Running {test_file}...')
                suite_name = test_file.stem
                timestamp = time.strftime("%m-%d_%H-%M-%S")
                results_dir = project_root / 'results' / f"{suite_name}_{timestamp}"
                os.makedirs(results_dir, exist_ok=True)
                result = subprocess.run([
                                'robot',
                                '--outputdir', str(results_dir),
                                '--include', tag,
                                str(test_file)
                    ]   + sys.argv[1:])
                if result.returncode != 0:
                    print(f'Error occurred while running {test_file}')
   

if __name__ == "__main__":
    run_SelectedTest('BS-Integration')