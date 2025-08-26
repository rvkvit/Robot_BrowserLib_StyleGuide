import os
import subprocess
import sys
import time
from openpyxl import load_workbook
from pathlib import Path

os_name = os.name

def run_SelectedTest():  
    if os_name == 'nt':  # Windows
       filepath = os.path.join(os.path.dirname(os.getcwd()), r'datafiles\scenarios.xlsx')
    else:  # Linux/Mac
        filepath = os.path.join(os.getcwd(), r'Template/datafiles/scenarios.xlsx')
    print(filepath)    
    workbook = load_workbook(filepath)
    sheet = workbook['Scenario']

    scenario_list = []
    modified_list = []

    for row in sheet.iter_rows(min_row=None, values_only=True):  # Skipping the header row
        Scenerio, Run_Status = row  # Unpack the row values
        if Run_Status and str(Run_Status).lower() == 'yes':  # Check if status is 'yes'
            scenario_list.append(Scenerio)  # Add to the list

    if not scenario_list: #check my list has any value
        print("No Scenerio marked to run") 

    for file in scenario_list:
            modified_list.append(file + '.robot')
           
    for file in modified_list:
            if file.endswith('.robot'):
                if os_name == 'nt':  # Windows
                    test_file = os.path.join(os.path.dirname(os.getcwd()), f'tests\{file}')
                else:  # Linux/Mac
                    test_file = os.path.join(os.getcwd(), f'Template/tests/{file}')
                print(f'Running {test_file}...')
                suite_name = os.path.splitext(os.path.basename(test_file))[0]
                timestamp = time.strftime("%m-%d_%H-%M-%S")
                project_dir = Path(__file__).resolve().parent.parent
                results_dir = project_dir/f"results/{suite_name}_{timestamp}"
                os.makedirs(results_dir, exist_ok=True)
                result = subprocess.run(['robot','--outputdir', results_dir ,test_file] + sys.argv[1:])
                if result.returncode != 0:
                    print(f'Error occurred while running {test_file}')
   

if __name__ == "__main__":
    run_SelectedTest()